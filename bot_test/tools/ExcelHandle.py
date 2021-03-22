#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/9/1 14:00
# @Author  : Smile_Mr
# @File    : ExcelHandle.py
import copy,time,os
import random

import xlwt,xlrd,os
from xlutils.copy import copy as xl_copy
from config.config import PROJECT_PATH
from openpyxl import load_workbook
from openpyxl.styles import Font

#todo:excel文件数据处理
class ExcelHandle:

    def __init__(self,project_name,file_name,sheet_name,*args,**kwargs):
        '''
        :param file_name: 文件名
        :param sheet_name: sheet
        '''
        self.project_name = project_name
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.book = xlrd.open_workbook(PROJECT_PATH + os.path.sep + 'data' + os.path.sep + project_name + os.path.sep + file_name)
        self.sheet = self.book.sheet_by_name(sheet_name)

        self.env_book = xlrd.open_workbook(PROJECT_PATH + os.path.sep + 'data' + os.path.sep  + 'environment.xlsx')
        self.env_sheet = self.env_book.sheet_by_name('environment')
        # self.write_book = xl_copy(self.book)
        self.write_book = load_workbook(PROJECT_PATH + os.path.sep+'data'+os.path.sep+ project_name + os.path.sep + file_name)
    #获取第一行列表
    def get_title(self):
        return self.sheet.row_values(0)
    #区分step1，step2，把它加入一个case
    def get_all_cases_two(self):
        cases = []
        titles =self.get_title()
        case = {}
        for row in range(1,self.sheet.nrows):
            if not case:
                case = {
                    'project_name':self.project_name,
                    'file_name' :self.file_name,
                    'sheet_name':self.sheet_name,
                    'steps':[],
                    "botActId":'test' + str(random.random())[-6:]
                }
            step = {}
            if row:
                for col, title in enumerate(titles):
                    step['row'] = row
                    step["botActId"] = case['botActId']
                    step[title] = self.sheet.cell_value(row,col)

            case['steps'].append(step)
            if row == self.sheet.nrows - 1 or self.sheet.cell_value(row + 1,titles.index('Steps')).upper() == "STEP1":
                cases.append(copy.deepcopy(case))
                case.clear()
        return cases
    #没有区分step1这个步骤
    def get_all_cases(self):
        cases = []
        titles = self.get_title()
        for row in range(1,self.sheet.nrows):
            case = {}
            if row:
                for col,title in enumerate(titles):
                    case[title] = self.sheet.cell_value(row,col)
                case['row'] = row
                case['file_name'] = self.file_name
                case['sheet_name'] = self.sheet_name
                cases.append(case)
        return cases
    #
    def write_back_result(self,row,actual,result=None,*args,**kwargs):
        pass_style = Font(color="00FF00")
        fail_style = Font(color="FF0000")
        style = pass_style if result =='Pass' else fail_style
        titles = self.get_title()
        col = titles.index('Actual')
        next_col = col + 1
        write_sheet = self.write_book[self.sheet_name]

        write_sheet.cell(row+1,col+1).value = actual
        write_sheet.cell(row+1,next_col+1).value = result

        write_sheet.cell(row+1,col+1).font = style
        write_sheet.cell(row+1,next_col+1).font= style
        self.save()

    def save(self):

        file_path = PROJECT_PATH + os.path.sep + 'results' + os.path.sep + self.project_name
        if not os.path.exists(file_path):
            os.mkdir(PROJECT_PATH + os.path.sep + 'results' + os.path.sep + self.project_name)
        result_file_path = file_path
        self.write_book.save(result_file_path + os.path.sep + self.file_name.replace('.xlsx', '.xls'))

    def get_environment(self):
        # env_datas = []
        env_data = {}
        for row in range(1,self.env_sheet.nrows):

            if self.env_sheet.cell_value(row,0) == self.project_name and self.env_sheet.cell_value(row,1) == self.file_name:
                env_data = {
                    'project_name':self.project_name,
                    'file_name':self.file_name,
                    'url_tran_data':[],
                    'playload_tran_data':[]
                }
                url_tran_data = self.env_sheet.cell_value(row,2)
                playload_tran_data = self.env_sheet.cell_value(row,3)
                # print(url_tran_data.split(','))
                env_data['url_tran_data'].extend(self.env_sheet.cell_value(row,2).split(','))
                env_data['playload_tran_data'].extend(self.env_sheet.cell_value(row,3).split(','))
                data = copy.deepcopy(env_data)
                # env_datas.append(data)
                # env_data.clear()
        return env_data
if __name__ == '__main__':
    print(ExcelHandle('KG','KG-8016.xlsx','Sheet1').get_environment())

